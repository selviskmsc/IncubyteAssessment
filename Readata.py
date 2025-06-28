import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl



try:
    file='data.xlsx'
    workbook=openpyxl.load_workbook(file)
    sheet=workbook["Sheet1"]

    rows=sheet.max_row
    cols=sheet.max_column


    firstName=print(sheet.cell(13,5).value)
    lastName=print(sheet.cell(18,5).value)
    eMail = print(sheet.cell(21, 5).value)
    password= print(sheet.cell(22, 5).value)
    confirmation = print(sheet.cell(23, 5).value)

    chromedriver_path = 'D:/chromedriver-win64/chromedriver.exe'

         #Create a Service object for the ChromeDriver
    service = Service(executable_path=chromedriver_path)

         #Initialize the Chrome WebDriver
    driver = webdriver.Chrome(service=service)

    driver.get("https://magento.softwaretestingboard.com/")
    driver.maximize_window()
    driver.implicitly_wait(10)


    driver.find_element(By.LINK_TEXT,"Create an Account").click()
    driver.execute_script("window.scrollBy(0,3000)")
    driver.find_element(By.CSS_SELECTOR,"input#firstname").send_keys(firstName)
    driver.find_element(By.CSS_SELECTOR,"input#lastname").send_keys(lastName)
    driver.find_element(By.CSS_SELECTOR,"input#email_address").send_keys(eMail)
    driver.find_element(By.CSS_SELECTOR,"input#password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR,"input#password-confirmation").send_keys(password)
    driver.find_element(By.XPATH,"//button[@class='action submit primary']").click()
    time.sleep(5)

except FileNotFoundError as fe:
    print(f"Error: The file was not found in the current directory. {fe}")
except Exception as e:
    print(f"An error occurred: {e}")
finally:
    driver.close()









